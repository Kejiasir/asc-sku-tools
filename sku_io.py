#!/usr/bin/env python3
"""Import and export subscription SKUs as JSON or Excel with a shared field set."""

from __future__ import annotations

import json
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from create_subscriptions import (
    ASCError,
    INTRO_DURATIONS,
    INTRO_MODES,
    PAYG_DURATION_BY_PERIOD,
    coerce_paid_intro_duration,
    intro_from_dict,
    normalize_period,
    parse_money,
)

EXCHANGE_FORMAT = "asc-sku.subscriptions"
EXCHANGE_VERSION = 1

COLUMNS: tuple[tuple[str, str], ...] = (
    ("group_level", "级别"),
    ("reference_name", "参考名称"),
    ("product_id", "产品 ID"),
    ("period", "持续时间"),
    ("usd_price", "美区价格"),
    ("intro_mode", "推介优惠类型"),
    ("intro_duration", "推介持续时间"),
    ("intro_periods", "推介期数"),
    ("intro_usd_price", "推介价格"),
    ("review_note", "审核备注"),
    ("review_screenshot", "审核截图"),
    ("state", "状态"),
)

HEADER_ALIASES = {
    key: key
    for key, _title in COLUMNS
}
HEADER_ALIASES.update(
    {
        title: key
        for key, title in COLUMNS
    }
)
HEADER_ALIASES.update(
    {
        "productid": "product_id",
        "产品id": "product_id",
        "referencename": "reference_name",
        "订阅期限": "period",
        "周期": "period",
        "美区价": "usd_price",
        "价格": "usd_price",
        "intro": "intro_mode",
        "推介促销优惠": "intro_mode",
        "推介促销优惠类型": "intro_mode",
        "优惠类型": "intro_mode",
        "持续时间": "period",
        "推介时长": "intro_duration",
        "number_of_periods": "intro_periods",
        "期数": "intro_periods",
        "intro_price": "intro_usd_price",
        "intro.usd_price": "intro_usd_price",
        "level": "group_level",
        "screenshot": "review_screenshot",
        "审核截图路径": "review_screenshot",
    }
)

PERIOD_LABELS = {
    "ONE_WEEK": "1 周",
    "ONE_MONTH": "1 个月",
    "TWO_MONTHS": "2 个月",
    "THREE_MONTHS": "3 个月",
    "SIX_MONTHS": "6 个月",
    "ONE_YEAR": "1 年",
}
PERIOD_IMPORT = {
    "1周": "ONE_WEEK",
    "一周": "ONE_WEEK",
    "1个月": "ONE_MONTH",
    "一个月": "ONE_MONTH",
    "2个月": "TWO_MONTHS",
    "两个月": "TWO_MONTHS",
    "3个月": "THREE_MONTHS",
    "一季度": "THREE_MONTHS",
    "6个月": "SIX_MONTHS",
    "半年": "SIX_MONTHS",
    "1年": "ONE_YEAR",
    "一年": "ONE_YEAR",
}

INTRO_MODE_LABELS = {
    "": "无",
    "PAY_AS_YOU_GO": "随用随付",
    "PAY_UP_FRONT": "提前支付",
    "FREE_TRIAL": "免费",
}
INTRO_MODE_IMPORT = {
    "无": "",
    "没有": "",
    "none": "",
    "随用随付": "PAY_AS_YOU_GO",
    "提前支付": "PAY_UP_FRONT",
    "免费": "FREE_TRIAL",
    "免费试用": "FREE_TRIAL",
}

DURATION_LABELS = {
    "THREE_DAYS": "3 天",
    "ONE_WEEK": "1 周",
    "TWO_WEEKS": "2 周",
    "ONE_MONTH": "1 个月",
    "TWO_MONTHS": "2 个月",
    "THREE_MONTHS": "3 个月",
    "SIX_MONTHS": "6 个月",
    "ONE_YEAR": "1 年",
}
DURATION_IMPORT = {
    "3天": "THREE_DAYS",
    "三天": "THREE_DAYS",
    "1周": "ONE_WEEK",
    "一周": "ONE_WEEK",
    "2周": "TWO_WEEKS",
    "两周": "TWO_WEEKS",
    "1个月": "ONE_MONTH",
    "一个月": "ONE_MONTH",
    "2个月": "TWO_MONTHS",
    "两个月": "TWO_MONTHS",
    "3个月": "THREE_MONTHS",
    "6个月": "SIX_MONTHS",
    "1年": "ONE_YEAR",
    "一年": "ONE_YEAR",
}

TEMPLATE_SKUS: list[dict[str, Any]] = [
    {
        "product_id": "dj.week.1.0",
        "reference_name": "vip-week-1.0-无试用",
        "period": "ONE_WEEK",
        "usd_price": "6.99",
        "group_level": 1,
        "review_note": "周订阅，无推介优惠",
    },
    {
        "product_id": "dj.month.1.1",
        "reference_name": "vip-month-1.1-前三天免费",
        "period": "ONE_MONTH",
        "usd_price": "29.99",
        "group_level": 2,
        "intro": {"mode": "FREE_TRIAL", "duration": "THREE_DAYS", "number_of_periods": 1},
        "review_note": "月订阅，前三天免费",
    },
    {
        "product_id": "dj.month.1.2",
        "reference_name": "vip-month-1.2-随用随付",
        "period": "ONE_MONTH",
        "usd_price": "29.99",
        "group_level": 3,
        "intro": {
            "mode": "PAY_AS_YOU_GO",
            "duration": "ONE_MONTH",
            "number_of_periods": 2,
            "usd_price": "0.99",
        },
    },
]

GUIDE_LINES = [
    "这是 ASC SKU 批量导入模板。填好「SKU」表后，在工具里点「导入」，再点「发布到 ASC」。",
    "",
    "必填：产品 ID、参考名称、持续时间、美区价格。",
    "选填：级别、推介优惠、审核备注、审核截图。状态仅用于导出对照，导入后不会写入 Apple。",
    "",
    "持续时间可填：1 周 / 1 个月 / 2 个月 / 3 个月 / 6 个月 / 1 年，或 ONE_WEEK、ONE_MONTH 等英文常量。",
    "推介优惠类型可填：无 / 随用随付 / 提前支付 / 免费。",
    "免费、提前支付：填「推介持续时间」（3 天、1 周、1 个月…）。",
    "随用随付：填「推介期数」1–12，持续时间必须与订阅期限相同；「推介价格」必填。",
    "提前支付 / 随用随付必须填推介价格；免费请留空。",
    "审核截图填本机 JPG/PNG 路径，可留空。",
    "",
    "JSON 与 Excel 字段相同。JSON 里推介优惠可以写成 intro 对象，也可以拆成与表格相同的扁平字段。",
    "导出文件可直接再导入。相同产品 ID 默认覆盖，也可选择清空后全量替换。",
]


def _norm_header(raw: Any) -> str:
    text = str(raw or "").strip().lower().replace(" ", "").replace("_", "")
    return text


def _header_key(raw: Any) -> str | None:
    text = str(raw or "").strip()
    if not text:
        return None
    if text in HEADER_ALIASES:
        return HEADER_ALIASES[text]
    compact = text.lower().replace(" ", "").replace("-", "_")
    if compact in HEADER_ALIASES:
        return HEADER_ALIASES[compact]
    folded = _norm_header(text)
    for alias, key in HEADER_ALIASES.items():
        if _norm_header(alias) == folded:
            return key
    return None


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        text = format(Decimal(str(value)), "f")
        if "." in text:
            text = text.rstrip("0").rstrip(".")
        return text
    return str(value).strip()


def _money_text(value: Any, field_name: str) -> str:
    text = _cell_text(value).replace("$", "").replace(",", "")
    if not text:
        raise ASCError(f"{field_name} 不能为空。")
    return str(parse_money(text, field_name))


def _optional_money(value: Any, field_name: str) -> str:
    text = _cell_text(value).replace("$", "").replace(",", "")
    if not text:
        return ""
    return str(parse_money(text, field_name))


def _parse_period(value: Any) -> str:
    text = _cell_text(value)
    if not text:
        raise ASCError("持续时间不能为空。")
    compact = text.replace(" ", "")
    if compact in PERIOD_IMPORT:
        return PERIOD_IMPORT[compact]
    return normalize_period(text)


def _parse_intro_mode(value: Any) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    mapped = INTRO_MODE_IMPORT.get(text) or INTRO_MODE_IMPORT.get(text.lower())
    if mapped is not None:
        return mapped
    mode = text.strip().upper().replace("-", "_").replace(" ", "")
    if mode in INTRO_MODES:
        return mode
    raise ASCError(f"不支持的推介优惠类型：{text}")


def _parse_duration(value: Any) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    compact = text.replace(" ", "")
    if compact in DURATION_IMPORT:
        return DURATION_IMPORT[compact]
    duration = text.strip().upper().replace("-", "_").replace(" ", "")
    if duration in INTRO_DURATIONS:
        return duration
    raise ASCError(f"不支持的推介持续时间：{text}")


def _parse_level(value: Any) -> int | None:
    text = _cell_text(value)
    if not text:
        return None
    try:
        return int(Decimal(text))
    except (InvalidOperation, ValueError) as error:
        raise ASCError(f"级别必须是整数：{value}") from error


def _row_get(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
        for actual, value in row.items():
            if _header_key(actual) == key and value not in (None, ""):
                return value
    return None


def flatten_sku(sku: dict[str, Any]) -> dict[str, Any]:
    intro = sku.get("intro") or {}
    return {
        "group_level": sku.get("group_level"),
        "reference_name": sku.get("reference_name", ""),
        "product_id": sku.get("product_id", ""),
        "period": sku.get("period", ""),
        "usd_price": sku.get("usd_price", ""),
        "intro_mode": intro.get("mode", "") if intro else "",
        "intro_duration": intro.get("duration", "") if intro else "",
        "intro_periods": intro.get("number_of_periods", "") if intro else "",
        "intro_usd_price": intro.get("usd_price", "") if intro else "",
        "review_note": sku.get("review_note", ""),
        "review_screenshot": sku.get("review_screenshot", ""),
        "state": sku.get("state", ""),
    }


def normalize_imported_sku(row: dict[str, Any], *, row_label: str) -> dict[str, Any]:
    try:
        product_id = _cell_text(_row_get(row, "product_id"))
        reference_name = _cell_text(_row_get(row, "reference_name"))
        if not product_id or " " in product_id:
            raise ASCError("产品 ID 不能为空，也不能包含空格。")
        if not reference_name:
            raise ASCError("参考名称不能为空。")
        period = _parse_period(_row_get(row, "period"))
        sku: dict[str, Any] = {
            "product_id": product_id,
            "reference_name": reference_name,
            "period": period,
            "usd_price": _money_text(_row_get(row, "usd_price"), "usd_price"),
        }
        level = _parse_level(_row_get(row, "group_level"))
        if level is not None:
            sku["group_level"] = level
        nested = row.get("intro")
        mode = _parse_intro_mode(_row_get(row, "intro_mode") if not nested else (nested or {}).get("mode"))
        if nested and isinstance(nested, dict) and nested.get("mode"):
            intro = dict(nested)
            intro["mode"] = _parse_intro_mode(intro.get("mode"))
            if intro.get("duration"):
                intro["duration"] = _parse_duration(intro.get("duration"))
            if intro.get("usd_price") not in (None, ""):
                intro["usd_price"] = _optional_money(intro.get("usd_price"), "intro.usd_price")
            parsed = intro_from_dict(intro)
            parsed = coerce_paid_intro_duration(period, parsed)
            if parsed is not None:
                payload: dict[str, Any] = {
                    "mode": parsed.mode,
                    "duration": parsed.duration,
                    "number_of_periods": parsed.number_of_periods,
                }
                if parsed.usd_price:
                    payload["usd_price"] = parsed.usd_price
                sku["intro"] = payload
        elif mode:
            duration = _parse_duration(_row_get(row, "intro_duration"))
            periods_raw = _cell_text(_row_get(row, "intro_periods"))
            periods = int(Decimal(periods_raw)) if periods_raw else 1
            intro_dict: dict[str, Any] = {
                "mode": mode,
                "duration": duration or (PAYG_DURATION_BY_PERIOD.get(period, period) if mode == "PAY_AS_YOU_GO" else ""),
                "number_of_periods": periods,
            }
            intro_price = _optional_money(_row_get(row, "intro_usd_price"), "intro_usd_price")
            if intro_price:
                intro_dict["usd_price"] = intro_price
            parsed = intro_from_dict(intro_dict)
            parsed = coerce_paid_intro_duration(period, parsed)
            if parsed is not None:
                payload = {
                    "mode": parsed.mode,
                    "duration": parsed.duration,
                    "number_of_periods": parsed.number_of_periods,
                }
                if parsed.usd_price:
                    payload["usd_price"] = parsed.usd_price
                sku["intro"] = payload
        note = _cell_text(_row_get(row, "review_note"))
        if note:
            sku["review_note"] = note
        screenshot = _cell_text(_row_get(row, "review_screenshot"))
        if screenshot:
            sku["review_screenshot"] = screenshot
        state = _cell_text(_row_get(row, "state"))
        if state:
            sku["state"] = state
        return sku
    except ASCError as error:
        raise ASCError(f"{row_label}: {error}") from error
    except (InvalidOperation, ValueError) as error:
        raise ASCError(f"{row_label}: invalid number ({error})") from error


def _subscriptions_from_payload(payload: Any) -> list[Any]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        items = payload.get("subscriptions")
        if isinstance(items, list):
            return items
    raise ASCError("JSON 需要是 SKU 数组，或包含 subscriptions 列表的对象。")


def load_skus_from_json(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    rows = _subscriptions_from_payload(payload)
    return [normalize_imported_sku(item, row_label=f"第 {index} 条") for index, item in enumerate(rows, start=1)]


def _excel_sheet_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["SKU"] if "SKU" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ASCError("Excel 是空的。")
    headers = [_header_key(cell) for cell in rows[0]]
    if "product_id" not in headers:
        raise ASCError("Excel 第一行必须包含「产品 ID」表头。不要改模板表头。")
    items: list[dict[str, Any]] = []
    for values in rows[1:]:
        if values is None or all(cell in (None, "") for cell in values):
            continue
        item: dict[str, Any] = {}
        for key, cell in zip(headers, values):
            if key:
                item[key] = cell
        items.append(item)
    return items


def load_skus_from_excel(path: Path) -> list[dict[str, Any]]:
    rows = _excel_sheet_rows(path)
    return [normalize_imported_sku(item, row_label=f"第 {index + 1} 行") for index, item in enumerate(rows, start=1)]


def load_skus(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return load_skus_from_json(path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_skus_from_excel(path)
    raise ASCError("只支持 .json 或 .xlsx。")


def canonical_sku(sku: dict[str, Any]) -> dict[str, Any]:
    return normalize_imported_sku(sku, row_label=str(sku.get("product_id") or "SKU"))


def dump_skus_json(skus: list[dict[str, Any]], path: Path) -> None:
    payload = {
        "format": EXCHANGE_FORMAT,
        "version": EXCHANGE_VERSION,
        "subscriptions": [canonical_sku(item) for item in skus],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _write_excel_rows(path: Path, skus: list[dict[str, Any]], *, guide: bool) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SKU"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1C2028")
    wrap = Alignment(wrap_text=True, vertical="center")
    for index, (_key, title) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(1, index, title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for row_index, sku in enumerate(skus, start=2):
        flat = flatten_sku(canonical_sku(sku))
        for column_index, (key, _title) in enumerate(COLUMNS, start=1):
            value = flat.get(key, "")
            if key == "period" and value in PERIOD_LABELS:
                value = PERIOD_LABELS[value]
            elif key == "intro_mode":
                value = INTRO_MODE_LABELS.get(str(value), value)
            elif key == "intro_duration" and value in DURATION_LABELS:
                value = DURATION_LABELS[value]
            sheet.cell(row_index, column_index, value if value not in (None, "") else None)
    widths = [10, 28, 22, 14, 12, 16, 16, 12, 12, 24, 28, 16]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(skus) + 1, 2)}"
    period_list = ",".join(PERIOD_LABELS.values())
    mode_list = ",".join(INTRO_MODE_LABELS.values())
    duration_list = ",".join(DURATION_LABELS.values())
    last_row = max(len(skus) + 50, 50)
    for formula, column in (
        (f'"{period_list}"', 4),
        (f'"{mode_list}"', 6),
        (f'"{duration_list}"', 7),
    ):
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{get_column_letter(column)}2:{get_column_letter(column)}{last_row}")
    if guide:
        notes = workbook.create_sheet("说明")
        notes.column_dimensions["A"].width = 92
        for index, line in enumerate(GUIDE_LINES, start=1):
            cell = notes.cell(index, 1, line)
            cell.alignment = wrap
    workbook.save(path)


def dump_skus_excel(skus: list[dict[str, Any]], path: Path) -> None:
    _write_excel_rows(path, skus, guide=True)


def dump_skus(skus: list[dict[str, Any]], path: Path) -> None:
    suffix = path.suffix.lower()
    if suffix == ".json":
        dump_skus_json(skus, path)
        return
    if suffix in {".xlsx", ".xlsm"}:
        dump_skus_excel(skus, path)
        return
    raise ASCError("只支持导出 .json 或 .xlsx。")


def write_json_template(path: Path) -> None:
    dump_skus_json(TEMPLATE_SKUS, path)


def write_excel_template(path: Path) -> None:
    dump_skus_excel(TEMPLATE_SKUS, path)


def merge_skus(
    existing: list[dict[str, Any]],
    incoming: list[dict[str, Any]],
    *,
    replace: bool,
) -> tuple[list[dict[str, Any]], int, int]:
    if replace:
        merged = [canonical_sku(item) for item in incoming]
        merged.sort(key=lambda item: int(item.get("group_level") or 10**9))
        return merged, len(merged), 0
    by_id = {str(item.get("product_id", "")): dict(item) for item in existing if item.get("product_id")}
    added = 0
    updated = 0
    for item in incoming:
        sku = canonical_sku(item)
        product_id = sku["product_id"]
        if product_id in by_id:
            by_id[product_id] = sku
            updated += 1
        else:
            by_id[product_id] = sku
            added += 1
    merged = list(by_id.values())
    merged.sort(key=lambda item: int(item.get("group_level") or 10**9))
    return merged, added, updated
